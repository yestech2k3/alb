# Copyright 2021 VMware, Inc.
# SPDX-License-Identifier: Apache License 2.0
import copy
import logging
import os
from functools import reduce

import pandas
import re
import random
from pkg_resources import parse_version
import avi.migrationtools.f5_converter.converter_constants as conv_const

from xlsxwriter import Workbook
from openpyxl import load_workbook

from avi.migrationtools.avi_migration_utils import MigrationUtil


LOG = logging.getLogger(__name__)
csv_writer_dict_list = []

# Added variable for checking progress and get overall object.
ppcount = 0
ptotal_count = 0
global fully_migrated
fully_migrated = 0
used_pool_groups = {}
used_pool = {}


class NsxtConvUtil(MigrationUtil):
    STATIC_PORT_MAP = {
        "http": conv_const.HTTP_PORT,
        "https": conv_const.HTTPS_PORT,
        "ftp": conv_const.FTP_PORT,
        "smtp": conv_const.SMTP_PORT,
        "snmp": conv_const.SNMP_PORT,
        "telnet": conv_const.TELNET_PORT,
        "snmp-trap": conv_const.SNMP_TRAP_PORT,
        "ssh": conv_const.SSH_PORT,
        "xfer": conv_const.XFER_PORT,
        "pcsync-https": conv_const.PCSYNC_HTTPS_PORT,
        "macromedia-fcs": conv_const.MACROMEDIA_FCS_PORT,
        "imap": conv_const.IMAP_PORT,
        "pop3": conv_const.POP3_PORT,
        "any": None
    }

    def add_conv_status(self, nsxt_type, nsxt_sub_type, nsxt_id, conv_status,
                        avi_object=None, need_review=None):
        """
        Adds as status row in conversion status csv
        :param nsxt_type: Object type
        :param nsxt_sub_type: Object sub type
        :param nsxt_id: Name oconv_f object
        :param conv_status: dict of conversion status
        :param avi_object: Converted objectconverted avi object
        """
        global csv_writer_dict_list
        # Added space if nsxt_sub_type None for pivot table
        row = {
            'NsxT type': nsxt_type,
            'NsxT SubType': nsxt_sub_type if nsxt_sub_type else ' ',
            'NsxT ID': nsxt_id,
            'Status': conv_status.get('status', ''),
            'Skipped settings': str(conv_status.get('skipped', '')),
            'Indirect mapping': str(conv_status.get('indirect', '')),
            'Not Applicable': str(conv_status.get('na_list', '')),
            'Avi Object': str(avi_object)
        }
        csv_writer_dict_list.append(row)

    def add_status_row(self, nsxt_type, nsxt_sub_type, nsxt_id, status, avi_obj=None):
        """
        Adds as status row in conversion status csv
        :param nsxt_type: Object type
        :param nsxt_sub_type: Object sub type
        :param nsxt_id: Name of object
        :param status: conversion status
        :param avi_obj: Converted avi object
        """
        global csv_writer_dict_list
        # Added space if nsxt_sub_type None for pivot table
        row = {
            'NsxT type': nsxt_type,
            'NsxT SubType': nsxt_sub_type if nsxt_sub_type else ' ',
            'NsxT ID': nsxt_id,
            'Status': status
        }
        if avi_obj:
            row.update({
                'Avi Object': str(avi_obj)
            })
        csv_writer_dict_list.append(row)

    def add_complete_conv_status(self, output_dir, avi_config, report_name,
                                 vs_level_status):

        global csv_writer_dict_list
        global ptotal_count
        for status in conv_const.STATUS_LIST:
            status_list = [row for row in csv_writer_dict_list if
                           row['Status'] == status]
            print('%s: %s' % (status, len(status_list)))
        print("Writing Excel Sheet For Converted Configuration...")
        ptotal_count = ptotal_count + len(csv_writer_dict_list)
        if vs_level_status:
            self.vs_per_skipped_setting_for_references(avi_config)
            self.correct_vs_ref(avi_config)
        else:
            # Update the complexity level of VS as Basic or Advanced
            self.vs_complexity_level()
        self.write_status_report_and_pivot_table_in_xlsx(
            output_dir, report_name, vs_level_status)

    def get_port_by_protocol(self, protocol):
        """
        Instead of default ports for protocols nsxt config has protocol in
        destination value for Avi object need to conver it to port number
        :param protocol: protocol name
        :return: integer value for protocol
        """

        return self.STATIC_PORT_MAP.get(protocol, None)

    def update_pool_for_service_port(self, pool_list, pool_name, hm_list,
                                     sys_hm_list):
        rem_hm = []
        pool = [obj for obj in pool_list if obj['name'] == pool_name]
        if pool:
            pool[0]['use_service_port'] = True
            # Checking monitor ports if use_service_port is true
            if pool[0].get('health_monitor_refs'):
                for hm in pool[0]['health_monitor_refs']:
                    hm_name = self.get_name(hm)
                    hm_ob = [ob for ob in (hm_list + sys_hm_list) if
                             ob['name'] == hm_name]
                    if hm_ob and (not hm_ob[0].get('monitor_port')):
                        rem_hm.append(hm)
                        LOG.debug("Removing monitor reference of %s from pool"
                                  " %s as 'use_service_port' is true but "
                                  "monitor has no port", hm_name,
                                  pool_name)
                if rem_hm:
                    pool[0]['health_monitor_refs'] = [
                        h_monitor for h_monitor in pool[0]
                        ['health_monitor_refs'] if h_monitor not in rem_hm]

                    rem_hm = [self.get_name(hmonitor) for hmonitor in rem_hm]
                    csv_row = [cl for cl in csv_writer_dict_list if cl[
                        'NsxT type'] == 'pool' and self.get_tenant_ref(
                        cl['NsxT ID'])[1] == pool_name]
                    if csv_row:
                        if csv_row[0]['Skipped settings'] in ('[]', ''):
                            csv_row[0]['Skipped settings'] = str([{
                                'monitor': rem_hm}])
                        else:
                            init_val = eval(csv_row[0]['Skipped settings'])
                            if not isinstance(init_val, list):
                                init_val = [init_val]
                            mon_val = [
                                val['monitor'].extend(rem_hm) for val in
                                init_val if isinstance(val, dict) and
                                            'monitor' in val]
                            if bool(mon_val):
                                csv_row[0]['Skipped settings'] = str(init_val)
                            else:
                                init_val.append({'monitor': rem_hm})
                                csv_row[0]['Skipped settings'] = str(init_val)
                        csv_row[0]['Status'] = conv_const.STATUS_PARTIAL
                        csv_row[0]['Avi Object'] = str({'pools': pool})

    def write_status_report_and_pivot_table_in_xlsx(
            self, output_dir, report_name, vs_level_status):
        """
        This function defines that add status sheet and pivot table sheet in
        xlsx format
        :param output_dir: Path of output directory
        :param report_name: filename to write report
        :param vs_level_status: Flag to include VS wise detailed status or not
        :return: None
        """
        global ppcount
        global ptotal_count
        # List of fieldnames for headers
        if vs_level_status:
            fieldnames = ['NsxT type', 'NsxT SubType', 'NsxT ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable', 'Complexity Level',
                          'VS Reference', 'Overall skipped settings',
                          'Avi Object']
        else:
            fieldnames = ['NsxT type', 'NsxT SubType', 'NsxT ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable', 'Complexity Level', 'Avi Object']

        # xlsx workbook
        report_path = output_dir + os.path.sep + "%s-ConversionStatus.xlsx" % \
                      report_name
        status_wb = Workbook(report_path)
        # xlsx worksheet
        status_ws = status_wb.add_worksheet("Status Sheet")
        # Lock the first row of xls report.
        status_ws.freeze_panes(1, 0)
        first_row = 0
        for header in fieldnames:
            col = fieldnames.index(header)
            status_ws.write(first_row, col, header)
        row = 1
        for row_data in csv_writer_dict_list:
            ppcount += 1
            for _key, _value in row_data.items():
                col = fieldnames.index(_key)
                status_ws.write(row, col, _value)
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
            row += 1
        status_wb.close()
        # create dataframe for row list
        df = pandas.DataFrame(csv_writer_dict_list, columns=fieldnames)
        # create pivot table using pandas
        pivot_table = \
            pandas.pivot_table(df, index=["Status", "NsxT type", "NsxT SubType"],
                               values=[], aggfunc=[len], fill_value=0)
        # create dataframe for pivot table using pandas
        pivot_df = pandas.DataFrame(pivot_table)
        main_book = \
            load_workbook(report_path)
        main_writer = pandas.ExcelWriter(report_path, engine='openpyxl')
        main_writer.book = main_book
        # Add pivot table in Pivot sheet
        pivot_df.to_excel(main_writer, 'Pivot Sheet')
        main_writer.save()

    def vs_complexity_level(self):
        """
        This method calculate the complexity of vs.
        :return:
        """
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['NsxT type'] == 'virtual']
        for vs_csv_object in vs_csv_objects:
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)

    def vs_per_skipped_setting_for_references(self, avi_config):
        """
        This functions defines that Add the skipped setting per VS CSV row
        :param avi_config: this method use avi_config for checking vs skipped
        :return: None
        """

        # Get the count of vs fully migrated
        global fully_migrated
        global ptotal_count
        global ppcount
        fully_migrated = 0
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['NsxT type'] == 'virtualservice']
        # Get the list of csv rows which has profile as NsxT type
        profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ['applicationprofile'])
        ssl_profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ['sslprofile'])
        ssl_key_certificate_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ['ssl_key_and_certificate'])
        ptotal_count = ptotal_count + len(vs_csv_objects)
        for vs_csv_object in vs_csv_objects:
            ppcount += 1
            skipped_setting = {}
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)
            vs_ref = virtual_service['name']
            repls = ('[', ''), (']', '')
            # Get list of skipped setting attributes
            skipped_setting_csv = reduce(lambda a, kv: a.replace(*kv), repls,
                                         vs_csv_object['Skipped settings'])
            if skipped_setting_csv:
                skipped_setting['virtual_service'] = [skipped_setting_csv]
            # Get the skipped list for ssl key and cert
            if 'ssl_key_and_certificate_refs' in virtual_service:
                for ssl_key_and_certificate_ref in \
                        virtual_service['ssl_key_and_certificate_refs']:
                    ssl_key_cert = self.get_name(ssl_key_and_certificate_ref)
                    ssl_kc_skip = self.get_csv_skipped_list(
                        ssl_key_certificate_csv_list, ssl_key_cert, vs_ref,
                        field_key='ssl_cert_key')
                    if ssl_kc_skip:
                        skipped_setting['ssl cert key'] = {}
                        skipped_setting['ssl cert key']['name'] = ssl_key_cert
                        skipped_setting['ssl cert key'][
                            'skipped_list'] = ssl_kc_skip

            # Get the skipped list for ssl profile name.
            # Changed ssl profile name to ssl profile ref.
            if 'ssl_profile_ref' in virtual_service:
                name, skipped = self.get_ssl_profile_skipped(
                    ssl_profile_csv_list, virtual_service['ssl_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['ssl profile'] = {}
                    skipped_setting['ssl profile']['name'] = name
                    skipped_setting['ssl profile']['skipped_list'] = skipped
            # Get the skipped list for pool group.
            if 'pool_group_ref' in virtual_service:
                pool_group_name = self.get_name(
                    virtual_service['pool_group_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                pool_group_skipped_settings = self.get_pool_skipped_list(
                    avi_config, pool_group_name, csv_pool_rows,
                    csv_writer_dict_list, vs_ref, profile_csv_list)
                if pool_group_skipped_settings:
                    skipped_setting['Pool Group'] = pool_group_skipped_settings
            # Get the skipped list for pool.
            if 'pool_ref' in virtual_service:
                pool_skipped_settings = {'pools': []}
                pool_name = self.get_name(virtual_service['pool_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                self.get_skipped_pool(
                    avi_config, pool_name, csv_pool_rows, csv_writer_dict_list,
                    vs_ref, profile_csv_list, pool_skipped_settings)
                if pool_skipped_settings['pools']:
                    skipped_setting['Pool'] = pool_skipped_settings
            # Get the skipepd list for http policy.
            if 'http_policies' in virtual_service:
                policy_csv_list = self.get_csv_object_list(
                    csv_writer_dict_list, ['policy', 'profile'])
                for http_ref in virtual_service['http_policies']:
                    policy_set_name, skipped_list = self.get_policy_set_skipped(
                        policy_csv_list, http_ref['http_policy_set_ref'],
                        vs_ref)
                    if skipped_list:
                        skipped_setting['Httppolicy'] = {}
                        skipped_setting['Httppolicy']['name'] = policy_set_name
                        skipped_setting['Httppolicy'][
                            'skipped_list'] = skipped_list
                    # Get the http policy name
                    pool_csv_rows = \
                        self.get_csv_object_list(csv_writer_dict_list, ['pool'])
                    for each_http_policy in avi_config['HTTPPolicySet']:
                        if each_http_policy['name'] == policy_set_name and 'http_request_policy' in each_http_policy:
                            for http_req in each_http_policy[
                                'http_request_policy']['rules']:
                                if http_req.get('switching_action', {}):
                                    self.get_skip_pools_policy(
                                        policy_set_name, http_req,
                                        avi_config, pool_csv_rows, vs_ref,
                                        profile_csv_list, skipped_setting)

            # # Get the skipped list for application_profile_ref.
            if 'application_profile_ref' in virtual_service:
                name, skipped = self.get_application_profile_skipped(
                    profile_csv_list,
                    virtual_service['application_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Application profile'] = {}
                    skipped_setting['Application profile'][
                        'name'] = name
                    skipped_setting['Application profile'][
                        'skipped_list'] = skipped
            # # Get the skipped list for network profile ref.
            if 'network_profile_ref' in virtual_service and 'admin:System' \
                    not in virtual_service['network_profile_ref']:
                name, skipped = self.get_network_profile_skipped(
                    profile_csv_list, virtual_service['network_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Network profile'] = {}
                    skipped_setting['Network profile'][
                        'name'] = name
                    skipped_setting['Network profile'][
                        'skipped_list'] = skipped
            # Update overall skipped setting of VS csv row
            if skipped_setting:
                vs_csv_object.update(
                    {'Overall skipped settings': str(skipped_setting)})
            else:
                vs_csv_object.update(
                    {'Overall skipped settings': "FULLY MIGRATION"})
                fully_migrated += 1
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
        csv_objects = [row for row in csv_writer_dict_list
                       if row['Status'] in [
                           conv_const.STATUS_PARTIAL,
                           conv_const.STATUS_SUCCESSFUL]
                       and row['NsxT type'] != 'virtualservice']

        # Update the vs reference not in used if objects are not attached to
        # VS directly or indirectly
        for row in csv_objects:
            if 'VS Reference' not in row or row['VS Reference'] == '':
                row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def correct_vs_ref(self, avi_config):
        """
        This method corrects the reference of VS to different objects
        :param avi_config: avi configuration dict
        :return:
        """
        global csv_writer_dict_list
        avi_graph = self.make_graph(avi_config)
        csv_dict_sub = [row for row in csv_writer_dict_list if row[
            'NsxT type'] != 'virtualservice' and row['Status'] in
                        (conv_const.STATUS_PARTIAL,
                         conv_const.STATUS_SUCCESSFUL)]
        for dict_row in csv_dict_sub:
            obj = dict_row['Avi Object']
            vs = []
            if obj.startswith('{'):
                obj = eval(obj)
                for key in obj:
                    for objs in obj[key]:
                        self.add_vs_ref(objs, avi_graph, vs)
            elif obj.startswith('['):
                obj = eval(obj)
                for objs in obj:
                    for key in objs:
                        objval = objs[key]
                        self.add_vs_ref(objval, avi_graph, vs)
            if vs:
                dict_row['VS Reference'] = str(list(set(vs)))
            else:
                dict_row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def get_vs_ssl_profiles(self, profiles, avi_config, prefix,
                            merge_object_mapping, sys_dict, f5_config):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param prefix: prefix for objects
        :param merge_object_mapping: Merged object mappings
        :param sys_dict: System object dict
        :return: returns list of profile refs assigned to VS in avi config
        """
        # f5_profiles = f5_config.get("profile", {})
        vs_ssl_profile_names = []
        pool_ssl_profile_names = []
        if not profiles:
            return vs_ssl_profile_names, pool_ssl_profile_names
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for key in profiles.keys():
            # Called tenant ref to get object name.
            tenant, name = self.get_tenant_ref(key)
            if prefix:
                name = prefix + '-' + name
            ssl_profile_list = avi_config.get("SSLProfile", [])
            sys_ssl = sys_dict['SSLProfile']
            ssl_profiles = [ob for ob in sys_ssl if ob['name'] ==
                            merge_object_mapping['ssl_profile'].get(name)
                            ] or [obj for obj in ssl_profile_list
                                  if (obj['name'] == name or name in
                                      obj.get("dup_of", []))]
            if ssl_profiles:
                cert_name = ssl_profiles[0].get('cert_name', None)
                if not cert_name:
                    cert_name = name
                ssl_key_cert_list = avi_config.get("SSLKeyAndCertificate", [])
                sys_key_cert = sys_dict['SSLKeyAndCertificate']
                key_cert = [ob for ob in sys_key_cert if ob['name'] ==
                            merge_object_mapping['ssl_cert_key'].get(cert_name)
                            ] or [obj for obj in ssl_key_cert_list if
                                  (obj['name'] == cert_name or obj['name'] ==
                                   '%s-%s' % (cert_name, conv_const.PLACE_HOLDER_STR) or cert_name in
                                   obj.get("dup_of", []))]
                # key_cert = key_cert[0]['name'] if key_cert else None
                if key_cert:
                    key_cert = self.get_object_ref(
                        key_cert[0]['name'], 'sslkeyandcertificate',
                        tenant=self.get_name(key_cert[0]['tenant_ref']))
                profile = profiles[key]
                context = profile.get("context") if profile else None
                if (not context) and isinstance(profile, dict):
                    if 'serverside' in profile:
                        context = 'serverside'
                    elif 'clientside' in profile:
                        context = 'clientside'
                pki_list = avi_config.get("PKIProfile", [])
                syspki = sys_dict['PKIProfile']
                pki_profiles = [ob for ob in syspki if ob['name'] ==
                                merge_object_mapping['pki_profile'].get(
                                    name)] or \
                               [obj for obj in pki_list if
                                (obj['name'] == name or
                                 name in obj.get("dup_of", []))]
                pki_profile = pki_profiles[0]['name'] if pki_profiles else None
                mode = 'SSL_CLIENT_CERTIFICATE_NONE'
                if pki_profile:
                    mode = pki_profiles[0].pop('mode',
                                               'SSL_CLIENT_CERTIFICATE_NONE')
                    pki_profile = self.get_object_ref(
                        pki_profiles[0]["name"], 'pkiprofile',
                        tenant=(pki_profiles[0]['tenant_ref'])).split('name=')[-1]
                if context == "clientside":
                    ssl_prof_ref = self.get_object_ref(
                        ssl_profiles[0]["name"], 'sslprofile',
                        tenant=(ssl_profiles[0]['tenant_ref'])).split('name=')
                    vs_ssl_profile_names.append({"profile": ssl_prof_ref,
                                                 "cert": key_cert,
                                                 "pki": pki_profile,
                                                 'mode': mode})
                elif context == "serverside":
                    ssl_prof_ref = self.get_object_ref(
                        ssl_profiles[0]["name"], 'sslprofile',
                        tenant=(ssl_profiles[0]['tenant_ref'])).split('name=')
                    pool_ssl_profile_names.append(
                        {"profile": ssl_prof_ref, "cert": key_cert,
                         "pki": pki_profile, 'mode': mode})
        return vs_ssl_profile_names, pool_ssl_profile_names

    def get_service_obj(self, destination, avi_config, enable_ssl,
                        controller_version, tenant_name, cloud_name, prefix,
                        vs_name, input_vrf=None):
        """
        Checks port overlapping scenario for port value 0 in F5 config
        :param destination: IP and Port destination of VS
        :param avi_config: Dict of avi config
        :param enable_ssl: value to put in service objects
        :param controller_version: Version of controller
        :param tenant_name: Name of tenant
        :param cloud_name: Name of cloud
        :param prefix: name prefix
        :param vs_name: Name of VS
        :param input_vrf: Vrf context name
        :return: services_obj, ip_addr of vs and ref of vsvip
        """

        parts = destination.split(':')
        ip_addr = parts[0]
        ip_addr = ip_addr.strip()
        vrf = None
        # Removed unwanted string from ip address
        if '%' in ip_addr:
            ip_addr, vrf = ip_addr.split('%')
        # Added support to skip virtualservice with ip address any
        if ip_addr == 'any':
            LOG.debug("Skipped:VS with IP address: %s" % str(destination))
            return None, None, None, None
        # Added check for IP V4
        matches = re.findall('^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', ip_addr)
        if not matches or ip_addr == '0.0.0.0':
            LOG.warning(
                'Avi does not support IPv6 Generated random ipv4 for vs:'
                ' %s' % ip_addr)
            ip_addr = ".".join(map(str, (
                random.randint(0, 255) for _ in range(4))))
        port = parts[1] if len(parts) == 2 else conv_const.DEFAULT_PORT
        # Get the list of vs which shared the same vip
        if parse_version(controller_version) >= parse_version('17.1'):
            # vs_dup_ips = \
            #     [vs for vs in avi_config['VirtualService'] if
            #      vs['vip'][0]['ip_address']['addr'] ==
            #      ip_addr]
            vs_dup_ips = []
            for vs in avi_config['VirtualService']:
                vs_ip = vs['vsvip_ref'].split('name=')[1].split('-')[0]
                if ip_addr == vs_ip:
                    vs_dup_ips.append(vs)
        else:
            vs_dup_ips = \
                [vs for vs in avi_config['VirtualService'] if
                 vs['ip_address']['addr'] == ip_addr]

        if port == 'any':
            port = '0'
        if isinstance(port, str) and (not port.isdigit()):
            port = self.get_port_by_protocol(port)
        # Port is None then skip vs
        if not port:
            LOG.debug("Skipped:Port not supported %s" % str(parts[1]))
            return None, None, None, None
        if int(port) > 0:
            for vs in vs_dup_ips:
                service_updated = self.update_service(port, vs, enable_ssl)
                if service_updated == 'duplicate_ip_port':
                    LOG.debug('Skipped: Duplicate IP-Port for vs %s', vs_name)
                    return None, None, None, None
                if service_updated:
                    break
            services_obj = [{'port': port, 'enable_ssl': enable_ssl}]
        else:
            used_ports = []
            for vs in vs_dup_ips:
                for service in vs['services']:
                    if service.get('port_range_end', None):
                        used_ports.extend(range(
                            int(service['port']),
                            int(service['port_range_end']) + 1
                        ))
                    else:
                        used_ports.append(int(service['port']))
            if used_ports and min(used_ports) == 1 and max(used_ports) == 65535:
                LOG.debug('Skipped: Duplicate IP-Port for vs %s', vs_name)
                return None, None, None, None
            if used_ports:
                services_obj = []
                if conv_const.PORT_END not in used_ports:
                    used_ports.append(conv_const.PORT_END + 1)
                used_ports = sorted(used_ports, key=int)
                start = conv_const.PORT_START
                for i in range(len(used_ports)):
                    if start == used_ports[i]:
                        start += 1
                        continue
                    end = int(used_ports[i]) - 1
                    if end < start:
                        start += 1
                        continue
                    services_obj.append({'port': start,
                                         'port_range_end': end,
                                         'enable_ssl': enable_ssl})
                    start = int(used_ports[i]) + 1
            else:
                services_obj = [
                    {'port': 1, 'port_range_end': conv_const.PORT_END,
                     'enable_ssl': enable_ssl}]
        # Getting vrf ref
        if vrf:
            self.add_vrf(avi_config, vrf, cloud_name)

        vrf_config = avi_config['VrfContext']
        vrf_ref = self.get_vrf_context_ref(destination, vrf_config,
                                           'virtual service', vs_name,
                                           cloud_name)
        if input_vrf:
            vrf_ref = self.get_object_ref(input_vrf, 'vrfcontext',
                                          cloud_name=cloud_name)
        if not vrf_ref:
            vrf_ref = self.get_object_ref('global', 'vrfcontext',
                                          cloud_name=cloud_name)

        updated_vsvip_ref = None
        if parse_version(controller_version) >= parse_version('17.1'):
            vs_vip_name = self.create_update_vsvip(
                ip_addr, avi_config['VsVip'],
                self.get_object_ref(tenant_name, 'tenant'),
                self.get_object_ref(cloud_name, 'cloud', tenant=tenant_name),
                prefix,
                vrf_ref)
            if vs_vip_name == '':
                updated_vsvip_ref = ''
            else:
                updated_vsvip_ref = self.get_object_ref(vs_vip_name, 'vsvip',
                                                        tenant_name, cloud_name)
        return services_obj, ip_addr, updated_vsvip_ref, vrf_ref

    def update_service(self, port, vs, enable_ssl):
        """
        iterates over services of existing vs in converted list to update
        services for port overlapping scenario
        :param port: port for currant VS
        :param vs: VS from converted config list
        :param enable_ssl: value to put in service object
        :return: boolean if service is updated or not
        """
        service_updated = False
        vs_new_service = []
        for service in vs['services']:
            port_end = service.get('port_range_end', None)
            if not port_end and int(service['port']) == int(port):
                return 'duplicate_ip_port'
            if port_end and (service['port'] <= int(port) <= port_end):
                if port not in [conv_const.PORT_START, conv_const.PORT_END]:
                    if service['port'] == int(port) == port_end:
                        return 'duplicate_ip_port'
                    elif service['port'] == int(port):
                        service['port'] = int(port) + 1
                    elif service['port_range_end'] == int(port):
                        service['port_range_end'] = int(port) - 1
                    else:
                        new_port = int(port) + 1
                        new_end = service['port_range_end']
                        service['port_range_end'] = int(port) - 1
                        new_service = {'port': new_port,
                                       'port_range_end': new_end,
                                       'enable_ssl': enable_ssl}
                        vs_new_service.append(new_service)
                elif port == conv_const.PORT_START:
                    service['port'] = 2
                elif port == conv_const.PORT_END:
                    service['port_range_end'] = (conv_const.PORT_START - 1)
                service_updated = True
                break
        vs['services'].extend(vs_new_service)
        return service_updated

    def add_vrf(self, avi_config, vrf, cloud_ref):
        vrf_name = 'vrf-%s' % vrf
        vrf_list = avi_config['VrfContext']
        vrf_obj = [obj for obj in vrf_list if obj['name'] == vrf_name]
        if not vrf_obj:
            vrf_obj = {
                "name": vrf_name,
                "system_default": False,
                "cloud_ref": self.get_object_ref(cloud_ref, 'cloud'),
                "tenant_ref": self.get_object_ref('admin', 'tenant')
            }
            if vrf_name == 'global':
                vrf_obj['system_default'] = True
            vrf_list.append(vrf_obj)

    def create_update_vsvip(self, vip, vsvip_config, tenant_ref, cloud_ref,
                            prefix, vrf_ref):
        """
        This functions defines that create or update VSVIP object.
        :param vip: vip of VS
        :param vsvip_config: List of vs object
        :param tenant_ref: tenant reference
        :param cloud_ref: cloud reference
        :param prefix: Name prefix
        :param vrf_ref: VRF reference
        :return: None
        """

        name = vip + '-vsvip'
        # Added prefix for objects
        if prefix:
            name = '%s-%s' % (prefix, name)
        # Get the exsting vsvip object list if present
        vsvip = [vip_obj for vip_obj in vsvip_config if vip_obj['name'] == name
                 and vip_obj.get('vrf_context_ref') == vrf_ref]
        if vsvip:
            diff_ten = [vips for vips in vsvip if vips['tenant_ref'] !=
                        tenant_ref]
            if diff_ten:
                LOG.debug('VsVip %s is repeated with vrf %s but different '
                          'tenant %s', name, self.get_name(vrf_ref) if vrf_ref
                          else 'None', self.get_name(tenant_ref))
                name = ''
        # If VSVIP object not present then create new VSVIP object.
        else:
            vsvip_object = {
                "name": name,
                "tenant_ref": tenant_ref,
                "cloud_ref": cloud_ref,
                "vip": [
                    {
                        "vip_id": "0",
                        "ip_address": {
                            "type": "V4",
                            "addr": vip
                        }
                    }
                ],
            }
            if vrf_ref:
                vsvip_object["vrf_context_ref"] = vrf_ref
            vsvip_config.append(vsvip_object)

        return name

    def get_vrf_context_ref(self, f5_entity_mem, vrf_config, entity_string,
                            entity_name, cloud):
        """
        Searches for vrf context refs in converted pool config
        :param f5_entity_mem: f5 entity or object like pool
        :param vrf_config: converted vrf config
        :param entity_string: entity string
        :param entity_name: name of f5 entity
        :param cloud: name of the cloud
        :return: returns list of vrf refs assigned to entity in avi config
        """
        vrf_ref = None
        f5_entity_mem = ':' in f5_entity_mem and f5_entity_mem.split(':')[0] \
                        or f5_entity_mem if f5_entity_mem else None
        vrf = 'vrf-' + f5_entity_mem.split('%')[1] \
            if f5_entity_mem and '%' in f5_entity_mem else None
        vrf_obj = [obj for obj in vrf_config if vrf and obj["name"] == vrf]
        if vrf_obj:
            vrf_ref = self.get_object_ref(
                vrf_obj[0]['name'], 'vrfcontext', cloud_name=cloud)
        else:
            LOG.warning("VRF not found for %s %s" % (entity_string,
                                                     entity_name))
        return vrf_ref

    def get_vs_app_profiles(self, profiles, avi_config, tenant_ref, prefix,
                            oc_prof, enable_ssl, merge_object_mapping,
                            sys_dict):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param tenant_ref: Tenant referance
        :param prefix: prefix for objects
        :param oc_prof: one connect profile
        :param enable_ssl: VS ssl enabled flag
        :param merge_object_mapping: Merged object mappings
        :param sys_dict: System object dict

        :return: returns list of profile refs assigned to VS in avi config
        """
        app_profile_refs = []
        app_prof_conf = dict()
        app_profile_list = avi_config.get("ApplicationProfile", [])
        unsupported_profiles = avi_config.get('UnsupportedProfiles', [])
        sys_app = sys_dict['ApplicationProfile']
        if not profiles:
            profiles = {}
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for name in profiles.keys():
            # Called tenant ref to get object name.
            name = self.get_tenant_ref(name)[1]
            # Added prefix for objects
            if prefix:
                name = '%s-%s' % (prefix, name)
            app_profiles = [ob for ob in sys_app if ob['name'] ==
                            merge_object_mapping['app_profile'].get(name)] or [
                               obj for obj in app_profile_list if
                               (obj['name'] == name
                                or name in obj.get("dup_of", []))]
            if app_profiles:
                app_prof_name = app_profiles[0]['name']
                app_profile_refs.append(self.get_object_ref(
                    app_prof_name, 'applicationprofile',
                    tenant=(app_profiles[0]['tenant_ref']))).split('name=')

                if app_profiles[0].get('HTTPPolicySet', None):
                    app_prof_conf['policy_name'] = app_profiles[0]['HTTPPolicySet']
                if app_profiles[0].get('fallback_host', None):
                    app_prof_conf['f_host'] = app_profiles[0]['fallback_host']
                # prerequisite user need to create default auth profile
                if app_profiles[0].get('realm', None):
                    app_prof_conf['realm'] = {
                        "type": "HTTP_BASIC_AUTH",
                        "auth_profile_ref": self.get_object_ref(
                            'System-Default-Auth-Profile', 'authprofile',
                            tenant=self.get_name(
                                app_profiles[0]['tenant_ref'])),
                        "realm": app_profiles[0]['realm']
                    }

        if not app_profile_refs:
            not_supported = [key for key in profiles.keys() if
                             key in unsupported_profiles]
            if not_supported:
                LOG.warning(
                    'Profiles not supported by Avi : %s' % not_supported)
                return app_prof_conf
            if enable_ssl:
                app_profile_refs.append(
                    self.get_object_ref('System-SSL-Application',
                                        'applicationprofile', tenant='admin'))
                app_prof_conf['app_prof'] = app_profile_refs
                return app_prof_conf
            else:
                app_profile_refs.append(
                    self.get_object_ref('System-L4-Application',
                                        'applicationprofile', tenant='admin'))
                app_prof_conf['app_prof'] = app_profile_refs
                return app_prof_conf
            # Added prefix for objects
            if prefix:
                value = '%s-%s' % (prefix, value)
            default_app_profile = [ob for ob in sys_app if ob['name'] ==
                                   merge_object_mapping['app_profile'].get(
                                       value)] or [
                                      obj for obj in app_profile_list if
                                      (obj['name'] == value
                                       or value in obj.get("dup_of", []))]
            tenant = (default_app_profile[0]['tenant_ref']).split('name=') if \
                default_app_profile else '/api/tenant/?name=admin'
            app_profile_refs.append(
                self.get_object_ref(default_app_profile[0]['name'],
                                    'applicationprofile', tenant=tenant))
        app_prof_conf['app_prof'] = app_profile_refs
        return app_prof_conf

    def get_vs_ntwk_profiles(self, profiles, avi_config, prefix,
                             merge_object_mapping, sys_dict):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param prefix: prefix for objects
        :param merge_object_mapping: merged object mappings
        :param sys_dict: System object dict
        :return: returns list of profile refs assigned to VS in avi config
        """
        network_profile_names = []
        if not profiles:
            return []
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for name in profiles.keys():
            # Called tenant method to get object name
            tenant, name = self.get_tenant_ref(name)
            # Added prefix for objects
            if prefix:
                name = prefix + '-' + name
            ntwk_prof_lst = avi_config.get("NetworkProfile")
            sysnw = sys_dict['NetworkProfile']
            network_profiles = [ob for ob in sysnw if
                                ob['name'] == merge_object_mapping[
                                    'network_profile'].get(name)] or \
                               [obj for obj in ntwk_prof_lst if (
                                       obj['name'] == name or name in
                                       obj.get("dup_of", []))]
            if network_profiles:
                network_profile_ref = self.get_object_ref(
                    network_profiles[0]['name'], 'networkprofile',
                    tenant=(network_profiles[0]['tenant_ref'])).split('name=')
                network_profile_names.append(network_profile_ref)
        return network_profile_names

    def get_application_profile_skipped(self, profile_csv_list, app_profile_ref,
                                        vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param app_profile_ref: Reference of application profile
        :param vs_ref: Name of VS
        :return: application profile name and skipped sttribute list
        """

        app_profile_name = self.get_name(app_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, app_profile_name, vs_ref, field_key='application_http_profile')
        return app_profile_name, skipped_list

    def get_vs_ntwk_profiles(self, profiles, avi_config, prefix,
                             merge_object_mapping, sys_dict):
        """
        Searches for profile refs in converted profile config if not found
        creates default profiles
        :param profiles: profiles in f5 config assigned to VS
        :param avi_config: converted avi config
        :param prefix: prefix for objects
        :param merge_object_mapping: merged object mappings
        :param sys_dict: System object dict
        :return: returns list of profile refs assigned to VS in avi config
        """
        network_profile_names = []
        if not profiles:
            return []
        if isinstance(profiles, str):
            profiles = profiles.replace(" {}", "")
            profiles = {profiles: None}
        for name in profiles.keys():
            # Called tenant method to get object name
            tenant, name = self.get_tenant_ref(name)
            # Added prefix for objects
            if prefix:
                name = prefix + '-' + name
            ntwk_prof_lst = avi_config.get("NetworkProfile")
            sysnw = sys_dict['NetworkProfile']
            network_profiles = [ob for ob in sysnw if
                                ob['name'] == merge_object_mapping[
                                    'network_profile'].get(name)] or \
                               [obj for obj in ntwk_prof_lst if (
                                       obj['name'] == name or name in
                                       obj.get("dup_of", []))]
            if network_profiles:
                network_profile_ref = self.get_object_ref(
                    network_profiles[0]['name'], 'networkprofile',
                    tenant=(network_profiles[0]['tenant_ref'])).split('name=')
                network_profile_names.append(network_profile_ref)
        return network_profile_names

    def get_app_persistence_profile_skipped(self, csv_writer_dict_list,
                                            pool_object, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: List of csv rows
        :param pool_object: object of pool
        :param vs_ref: Name of VS
        :return: profile name and skipped attribute list
        """

        app_persistence_profile_name = self.get_name(
            pool_object['application_persistence_profile_ref'])
        csv_object = self.get_csv_object_list(csv_writer_dict_list,
                                              ['persistence'])
        skipped_list = self.get_csv_skipped_list(
            csv_object, app_persistence_profile_name, vs_ref,
            field_key='app_per_profile')
        return app_persistence_profile_name, skipped_list

    def get_ssl_profile_skipped(self, profile_csv_list, ssl_profile_ref,
                                vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param ssl_profile_ref: Reference of ssl profile
        :param vs_ref: Name of VS
        :return: ssl profile name and skipped sttribute list
        """

        ssl_profile_name = self.get_name(ssl_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, ssl_profile_name, vs_ref, field_key='ssl_profile')
        return ssl_profile_name, skipped_list

    def get_conv_status_by_obj_name(self, obj_name):
        conv_status_list = list(filter(lambda obj: obj.get("NsxT ID") == obj_name, csv_writer_dict_list))
        conv_status = dict()
        conv_status["status"] = conv_status_list[0]["Status"]
        conv_status['skipped'] = conv_status_list[0]["Skipped settings"]
        conv_status['indirect'] = conv_status_list[0]["Indirect mapping"]
        conv_status['na_list'] = conv_status_list[0]['Not Applicable']
        return conv_status

    def clone_pool_if_shared(self, ref, avi_config, vs_name, tenant, p_tenant,
                             persist_type, controller_version, app_prof_ref,is_pool_group_used,
                              cloud_name='Default-Cloud', prefix=None):
        """
        clones pool or pool group if its shard between multiple VS or partitions
        in Nsxt
        :param ref: reference of pool or pool group
        :param avi_config: Avi configuration cloned pool or pool groups to be
        added
        :param vs_name: Name of the vs to be added
        :param tenant: tenant name of vs
        :param p_tenant: tenant name of pool
        :param persist_type: persistence profile type
        :param controller_version:
        :param app_prof_ref: Application profile referance
        :param sysdict:
        :param cloud_name:
        :param prefix:
        :return:
        """
        is_pool_group = False
        pool_group_obj = None
        # Added prefix for objects
        if prefix:
            ref = prefix + '-' + ref
        # Search the pool or pool group with name in avi config for the same
        # tenant as VS
        pool_obj = [pool for pool in avi_config['Pool'] if pool['name'] == ref
                    and pool['tenant_ref'] == self.get_object_ref(tenant,
                                                                  'tenant')]
        pool_per_ref = pool_obj[0].get(
            'application_persistence_profile_ref') if pool_obj else None
        pool_per_name = self.get_name(pool_per_ref) if pool_per_ref else None
        pool_per_types = [obj['persistence_type'] for obj in (avi_config[
                                                                  'ApplicationPersistenceProfile']) if obj['name'] ==
                          pool_per_name] if pool_per_name else []
        pool_per_type = pool_per_types[0] if pool_per_types else None
        if not pool_obj:
            pool_group_obj = [pool for pool in avi_config['PoolGroup']
                              if pool['name'] == ref and
                              pool['tenant_ref'] == self.get_object_ref(
                    tenant, 'tenant')]
        if pool_group_obj:
            is_pool_group = True
        if p_tenant:
            shared_vs = [obj for obj in avi_config['VirtualService']
                         if obj.get("pool_ref", "") == self.get_object_ref(
                    ref, 'pool', tenant=p_tenant, cloud_name=cloud_name)]
            if not shared_vs:
                shared_vs = [obj for obj in avi_config['VirtualService']
                             if obj.get("pool_group_ref", "") ==
                             self.get_object_ref(
                                 ref, 'poolgroup', tenant=p_tenant,
                                 cloud_name=cloud_name)]
        else:
            shared_vs = [obj for obj in avi_config['VirtualService']
                         if obj.get("pool_ref", "") == self.get_object_ref(
                    ref, 'pool', tenant=tenant, cloud_name=cloud_name)]
            if not shared_vs:
                shared_vs = [obj for obj in avi_config['VirtualService']
                             if obj.get("pool_group_ref", "") ==
                             self.get_object_ref(
                                 ref, 'poolgroup', tenant=tenant,
                                 cloud_name=cloud_name)]
        if not tenant == p_tenant :
            if is_pool_group:
                ref = self.clone_pool_group(ref, vs_name, avi_config, True,
                                            tenant, cloud_name=cloud_name)
            else:
                ref = self.clone_pool(ref, vs_name, avi_config['Pool'],
                                      True, tenant)
        if is_pool_group and not shared_vs:
            if ref in is_pool_group_used.keys():
                if pool_group_obj[0].get('members'):
                    if pool_group_obj[0]['members'][0]['pool_ref'].split('cloud=')[-1] != cloud_name:
                        shared_vs = is_pool_group_used.get(ref)
                        is_pool_group = True

        if shared_vs:
            if is_pool_group:
                ref = self.clone_pool_group(ref, vs_name, avi_config, True,
                                            tenant, cloud_name=cloud_name)
            else:
                shared_appref = shared_vs[0].get('application_profile_ref')
                shared_apptype = None
                if shared_appref:
                    shared_appname = self.get_name(shared_appref)
                    shared_appobjs = [ob for ob in (avi_config[
                                                        'ApplicationProfile']) if ob['name'] ==
                                      shared_appname]
                    shared_appobj = shared_appobjs[0] if shared_appobjs else {}
                    shared_apptype = shared_appobj['type'] if shared_appobj \
                        else None
                app_prof_name = self.get_name(app_prof_ref)
                app_prof_objs = [appob for appob in (avi_config[
                                                         'ApplicationProfile']) if appob['name'] ==
                                 app_prof_name]
                app_prof_obj = app_prof_objs[0] if app_prof_objs else {}
                app_prof_type = app_prof_obj['type'] if app_prof_obj else None

                if self.is_pool_clone_criteria(
                        controller_version, app_prof_type, shared_apptype,
                        persist_type, pool_per_type, shared_appobj,
                        app_prof_obj):
                    LOG.debug('Cloned the pool %s for VS %s', ref, vs_name)
                    ref = self.clone_pool(ref, vs_name, avi_config['Pool'],
                                          True, tenant)
                else:
                    LOG.debug("Shared pool %s for VS %s", ref, vs_name)

        return ref, is_pool_group

    def is_pool_clone_criteria(self, controller_version, app_prof_type,
                               shared_apptype, persist_type, pool_per_type,
                               shared_appobj, app_prof_obj):
        if parse_version(controller_version) < parse_version(
           '17.1.6') or app_prof_type != 'APPLICATION_PROFILE_TYPE_HTTP' \
           or shared_apptype != app_prof_type or (
                persist_type != None and persist_type !=
                'PERSISTENCE_TYPE_HTTP_COOKIE') or (
                pool_per_type != None and pool_per_type !=
                'PERSISTENCE_TYPE_HTTP_COOKIE') or (
                shared_appobj.get('http_profile', {}).get(
                    'connection_multiplexing_enabled') != app_prof_obj.get(
                    'http_profile', {}).get('connection_multiplexing_enabled') or (
                shared_appobj.get('http_profile', {}).get(
                    'cache_config') != app_prof_obj.get(
                    'http_profile', {}).get('cache_config'))):
            return True
        else:
            return False

    def clone_pool_group(self, pool_group_name, clone_for, avi_config, is_vs,
                         tenant='admin', cloud_name='Default-Cloud'):
        """
        If pool is shared with other VS pool is cloned for other VS as Avi dose
        not support shared pools with new pool name as <pool_name>-<vs_name>
        :param pool_group_name: Name of the pool group to be cloned
        :param clone_for: Name of the object/entity for pool group to be cloned
        :param avi_config: new pool to be added to avi config
        :param is_vs: True if clone is called for VS
        :param tenant: if nsxt pool is shared across partition then coned for
        tenant
        :param cloud_name:
        :return: new pool group name
        """
        pg_ref = None
        new_pool_group = None
        for pool_group in avi_config['PoolGroup']:
            if pool_group["name"] == pool_group_name:
                new_pool_group = copy.deepcopy(pool_group)
                break
        if new_pool_group:
            if pool_group_name in used_pool_groups:
                used_pool_groups[pool_group_name] += 1
            else:
                used_pool_groups[pool_group_name] = 1
            LOG.debug('Cloning pool group for %s', clone_for)
            new_pool_group["name"] = '{}-{}'.format(
                pool_group_name, used_pool_groups[pool_group_name])
            pg_ref = new_pool_group["name"]
            new_pool_group["tenant_ref"] = self.get_object_ref(tenant, 'tenant')
            avi_config['PoolGroup'].append(new_pool_group)
            for member in new_pool_group['members']:
                pool_name = self.get_name(member['pool_ref'])
                pool_name = self.clone_pool(pool_name, clone_for,
                                            avi_config['Pool'], is_vs, tenant)
                member['pool_ref'] = self.get_object_ref(
                    pool_name, 'pool', tenant=tenant, cloud_name=cloud_name)
        return pg_ref

    def clone_pool(self, pool_name, clone_for, avi_pool_list, is_vs,
                   tenant=None):
        """
        If pool is shared with other VS pool is cloned for other VS as Avi dose
        not support shared pools with new pool name as <pool_name>-<vs_name>
        :param pool_name: Name of the pool to be cloned
        :param clone_for: Name of the VS for pool to be cloned
        :param avi_pool_list: new pool to be added to this list
        :param is_vs: True if this cloning is for VS
        :param tenant: if pool is shared across partition then coned for tenant
        :return: new pool object
        """
        LOG.debug("Cloning pool %s for %s " % (pool_name, clone_for))
        new_pool = None
        for pool in avi_pool_list:
            if pool["name"] == pool_name:
                new_pool = copy.deepcopy(pool)
                break
        if new_pool:
            if pool_name in used_pool:
                used_pool[pool_name] += 1
            else:
                used_pool[pool_name] = 1
            LOG.debug('Cloning Pool for %s', clone_for)
            new_pool["name"] = '{}-{}'.format(pool_name, used_pool[pool_name])
            if tenant:
                new_pool["tenant_ref"] = self.get_object_ref(tenant, 'tenant')
            if is_vs:
                # removing config added from VS config to pool
                new_pool["application_persistence_profile_ref"] = None
                new_pool["ssl_profile_ref"] = None
                new_pool["ssl_key_and_certificate_ref"] = None
                new_pool["pki_profile_ref"] = None
                if new_pool.get('placement_networks'):
                    del new_pool['placement_networks']
            avi_pool_list.append(new_pool)
            pool_ref = new_pool["name"]
            LOG.debug("Cloned pool successfully %s for %s " % (
                new_pool["name"], clone_for))
            return pool_ref