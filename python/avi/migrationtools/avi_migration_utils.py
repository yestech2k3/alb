# Copyright 2021 VMware, Inc.
# SPDX-License-Identifier: Apache License 2.0

import pandas
import argparse
import ast
import copy
import getpass
import logging
import os
import random
import string
from functools import reduce
from urllib.parse import urlparse, parse_qs
from datetime import datetime
from socket import gethostname

import networkx as nx
import pexpect
import yaml
from OpenSSL import crypto
from openpyxl import load_workbook, Workbook

import avi.migrationtools.f5_converter.converter_constants as conv_const

LOG = logging.getLogger(__name__)
csv_writer_dict_list = []
tenants = []
ran_str = ''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase
                                + string.digits) for _ in range(3))
warning_count = 0
error_count = 0


def set_update_count():
    global warning_count, error_count
    warning_count = 0
    error_count = 0


def update_count(type='warning'):
    global warning_count, error_count
    if type == 'warning':
        warning_count += 1
    elif type == 'error':
        error_count += 1


def get_count(type='None'):
    if type == 'warning':
        return warning_count
    elif type == 'error':
        return error_count
    return {'warning': warning_count, 'error': error_count}


class PasswordPromptAction(argparse.Action):
    def __call__(self, parser, args, values, option_string=None):
        if values:
            setattr(args, self.dest, values)
        else:
            setattr(args, self.dest, getpass.getpass(prompt=self.dest))


class MigrationUtil(object):

    def add_conv_status(self, **args):
        pass

    def add_status_row(self, **args):
        pass

    def get_conv_status(self, skipped, indirect_list, ignore_dict, input_object,
                        user_ignore=None, na_list=None):
        """
        Update skipped list for conversion status
        :param skipped: All skipped attributes after conversion
        :param indirect_list: List of attrs to be mapped as indirect mapping
        :param ignore_dict: Dict of default values for column skipped for defaults
        :param input_object: Currant input object
        :param user_ignore: List of attributes user wants not to be shown in skipped
        :param na_list: List of attributes marked as not applicable
        :return: Conversion status dict
        """
        conv_status = dict()
        user_ignore = [] if not user_ignore else user_ignore
        na_list = [] if not na_list else na_list

        conv_status['user_ignore'] = [val for val in skipped if
                                      val in user_ignore]
        skipped = [attr for attr in skipped if attr not in user_ignore]

        conv_status['indirect'] = [val for val in skipped if
                                   val in indirect_list]
        skipped = [attr for attr in skipped if attr not in indirect_list]

        conv_status['na_list'] = [val for val in skipped if val in na_list]
        skipped = [attr for attr in skipped if attr not in na_list]

        default_skip = []
        for key in ignore_dict.keys():
            val = input_object.get(key)
            default_val = ignore_dict.get(key)
            if key in skipped and val == default_val:
                default_skip.append(key)
        if default_skip:
            skipped = [attr for attr in skipped if attr not in default_skip]

        conv_status['skipped'] = skipped
        conv_status['default_skip'] = default_skip
        if skipped:
            status = conv_const.STATUS_PARTIAL
        else:
            status = conv_const.STATUS_SUCCESSFUL
        conv_status['status'] = status
        return conv_status

    def get_tenant_ref(self, name):
        tenant = 'admin'
        if name and name.startswith('/'):
            parts = name.split('/', 2)
            tenant = parts[1]
            if not parts[2]:
                LOG.warning('Invalid tenant ref : %s' % name)
            name = parts[2]
        elif name and '/' in name:
            parts = name.split('/')
            # Changed the index to get the tenant and name in case of
            # prefixed name
            tenant = parts[-2]
            name = parts[-1]
        if tenant.lower() == 'common':
            tenant = 'admin'
        if '/' in name:
            name = name.split('/')[1]
        if ' ' in tenant:
            tenant = tenant.split(' ')[-1]
        tenant.strip()
        return tenant, name

    def create_self_signed_cert(self):
        # create a key pair
        key = crypto.PKey()
        key.generate_key(crypto.TYPE_RSA, 2048)

        # create a self-signed cert
        cert = crypto.X509()
        cert.get_subject().C = "US"
        cert.get_subject().O = "Avi Networks"
        cert.get_subject().CN = gethostname()
        cert.set_serial_number(1000)
        cert.gmtime_adj_notBefore(0)
        cert.gmtime_adj_notAfter(10 * 365 * 24 * 60 * 60)
        cert.set_issuer(cert.get_subject())
        cert.set_pubkey(key)
        cert.sign(key, 'sha256')
        cert = crypto.dump_certificate(crypto.FILETYPE_PEM, cert)
        key = crypto.dump_privatekey(crypto.FILETYPE_PEM, key)
        return key, cert

    def is_certificate_key_protected(self, key_file):
        """
        This functions defines that whether key is passphrase protected or not
        :param key_file: Path of key file
        :return: Return True if key is passphrase protected else return False
        """
        try:
            child = pexpect.spawn(
                'openssl rsa -in %s -check -noout' % key_file)
            # Expect for enter pass phrase if key is protected else it will raise
            # an exception
            child.expect('Enter pass phrase for')
            update_count('warning')
            return True
        except Exception as e:
            return False

    def get_pool_skipped_list(self, avi_config, pool_group_name, csv_pool_rows,
                              csv_writer_dict_list, vs_ref, profile_csv_list):
        """
        This method is used for getting pool skipped list.
        :param avi_config: AVI dict
        :param pool_group_name: Name of Pool group
        :param csv_pool_rows: List of pool(NsxT type) csv rows
        :param csv_writer_dict_list: List of nsxt csv rows
        :param vs_ref: Name of VS
        :param profile_csv_list: List of profile(NsxT type) csv rows
        :return:
        """
        pool_group_objects = list(filter(lambda pg: pg["name"] == pool_group_name, avi_config['PoolGroup']))
        pool_members = pool_group_objects[0]['members']
        skipped_setting = {
            'pools': []
        }
        for pool_member in pool_members:
            pool_name = self.get_name(pool_member['pool_ref'])
            self.get_skipped_pool(
                avi_config, pool_name, csv_pool_rows, csv_writer_dict_list,
                vs_ref, profile_csv_list, skipped_setting)
        if skipped_setting['pools']:
            return skipped_setting
        return None

    def get_skipped_pool(self, avi_config, pool_name, pool_csv_rows,
                         csv_writer_dict_list, vs_ref, profile_csv_list,
                         skipped_setting):
        """
        This method get the skipped list for pool by going over the
        references attached to it
        :param avi_config: Converted Avi configuration
        :param pool_name: name of the pool
        :param pool_csv_rows:
        :param csv_writer_dict_list: Result report dict
        :param vs_ref: VS reference
        :param profile_csv_list:
        :param skipped_setting: User defined skipped settings
        :return: skipped setting for pool
        """
        pool_skipped_setting = {}
        skipped_list = self.get_pool_skipped(pool_csv_rows, pool_name, vs_ref)
        pool_object = [pool for pool in avi_config["Pool"]
                       if pool['name'] == pool_name]
        if skipped_list:
            pool_skipped_setting['pool_name'] = pool_name
            pool_skipped_setting['pool_skipped_list'] = skipped_list
        if pool_object:
            if 'health_monitor_refs' in pool_object[0]:
                health_monitor_skipped_setting = []
                for health_monitor_ref in pool_object[0]['health_monitor_refs']:
                    health_monitor_ref = self.get_name(health_monitor_ref)
                    monitor_csv_object = self.get_csv_object_list(
                        csv_writer_dict_list, ['monitor'])
                    skipped_list = self.get_csv_skipped_list(
                        monitor_csv_object, health_monitor_ref, vs_ref,
                        field_key='health_monitor')
                    if skipped_list:
                        health_monitor_skipped_setting.append(
                            {'health_monitor_name': health_monitor_ref,
                             'monitor_skipped_list': skipped_list})
                if health_monitor_skipped_setting:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['health_monitor'] = \
                        health_monitor_skipped_setting
            if pool_object[0].get('ssl_key_and_certificate_ref', None):
                ssl_key_cert = self.get_name(
                    pool_object[0]['ssl_key_and_certificate_ref'])
                LOG.debug('[SslKeyAndCertificate] certificate {}'.format(ssl_key_cert))
                sslkc_skip = self.get_csv_skipped_list(
                    profile_csv_list, ssl_key_cert, vs_ref,
                    field_key='ssl_cert_key')
                if sslkc_skip:
                    LOG.debug('[SslKeyAndCertificate] Skipped Attribute {}'.format(sslkc_skip))
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['ssl_key_and_certificate'] = sslkc_skip
            else:
                LOG.info('Ssl key and certificate ref is not found')
            if pool_object[0].get('ssl_profile_ref', None):
                name, skipped = self.get_ssl_profile_skipped(
                    profile_csv_list, pool_object[0]['ssl_profile_ref'], vs_ref)
                if skipped:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['ssl profile'] = {}
                    pool_skipped_setting['ssl profile']['name'] = name
                    pool_skipped_setting['ssl profile'][
                        'skipped_list'] = skipped

            if 'application_persistence_profile_ref' in pool_object[0] and \
                    pool_object[0]['application_persistence_profile_ref']:
                name, skipped = self.get_app_persistence_profile_skipped(
                    csv_writer_dict_list, pool_object[0], vs_ref)
                if skipped:
                    pool_skipped_setting['pool_name'] = pool_name
                    pool_skipped_setting['Application Persistence profile'] = {}
                    pool_skipped_setting['Application Persistence profile'][
                        'name'] = name
                    pool_skipped_setting['Application Persistence profile'][
                        'skipped_list'] = skipped

            if pool_skipped_setting:
                skipped_setting['pools'].append(pool_skipped_setting)
        else:
            LOG.debug('[PoolObject] Not Found for pool {}'.format(pool_name))

    def get_pool_skipped(self, csv_objects, pool_name, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_objects: CSV row of object from xlsx report
        :param pool_name: Name of pool
        :param vs_ref: Name of VS
        :return: Skipped list of csv row
        """

        for csv_object in csv_objects:
            avi_object = self.format_string_to_json(csv_object['Avi Object'])
            if 'pools' in avi_object:
                pool_object = [pool for pool in avi_object['pools']
                               if pool['name'] == pool_name]
                if pool_object:
                    return self.get_and_update_csv_row(csv_object, vs_ref)

    def get_csv_skipped_list(self, csv_objects, name_of_object, vs_ref,
                             field_key=None):
        """
        This method is used for getting skipped list from vs.
        :param csv_objects: CSV row of object from xlsx report
        :param name_of_object: Name of object
        :param vs_ref: Name of VS
        :param field_key: Key fromm avi json which is specific for object type
        :return: Return skipped attribute list
        """
        for csv_object in csv_objects:
            avi_objects = self.format_string_to_json(csv_object['Avi Object'])
            if isinstance(avi_objects, dict):
                avi_objects = [avi_objects]
            if not avi_objects:
                avi_objects = []
            for avi_object_json in avi_objects:
                object_found = False
                if field_key:
                    if field_key in avi_object_json and 'Duplicate' not in \
                            avi_object_json[field_key] and (
                            avi_object_json[field_key]['name'] ==
                            name_of_object):
                        object_found = True
                else:
                    if avi_object_json.get('name') and \
                            avi_object_json['name'] == name_of_object:
                        object_found = True

                if object_found:
                    return self.get_and_update_csv_row(csv_object, vs_ref)

    def get_and_update_csv_row(self, csv_object, vs_ref):
        """
        This function defines that update csv row.
        :param csv_object: csv object
        :param vs_ref: Name of VS
        :return: Skipped attribute list
        """

        if 'VS Reference' in csv_object and \
                vs_ref not in csv_object['VS Reference']:
            csv_object['VS Reference'] += ',' + vs_ref
        else:
            csv_object['VS Reference'] = vs_ref
        repls = ('[', ''), (']', '')
        skipped_setting_csv = reduce(
            lambda a, kv: a.replace(*kv), repls, csv_object['Skipped settings'])
        if skipped_setting_csv:
            return [skipped_setting_csv]

    def get_csv_object_list(self, csv_writer_dict_list, command_list):
        """
        This method is used for getting csv object
        :param csv_writer_dict_list: CSV row of object from xlsx report
        :param command_list: List of netscaler commands
        :return: List of CSV rows
        """

        csv_object = [row for row in csv_writer_dict_list if
                      row['Status'] in [conv_const.STATUS_PARTIAL,
                                        conv_const.STATUS_SUCCESSFUL] and
                      '->' not in row['Avi Object'] and
                      row['NsxT type'] in command_list]
        return csv_object

    def format_string_to_json(self, avi_string):
        """
        This function defines that it convert string into json format to
        convert into dict
        :param avi_string: string to be converted
        :return: Return converted string
        """
        avi_string = avi_string.split('__/__')[0]
        return ast.literal_eval(avi_string)

    def get_app_persistence_profile_skipped(self, csv_writer_dict_list,
                                            pool_object, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param csv_writer_dict_list: List of csv rows
        :param pool_object: object of pool
        :param vs_ref: Name of VS
        :return: profile name and skipped attribute list
        """

        app_persistence_profile_name = self.get_name(
            pool_object['application_persistence_profile_ref'])
        csv_object = self.get_csv_object_list(csv_writer_dict_list,
                                              ['persistence'])
        skipped_list = self.get_csv_skipped_list(
            csv_object, app_persistence_profile_name, vs_ref,
            field_key='app_per_profile')
        return app_persistence_profile_name, skipped_list

    def get_ssl_profile_skipped(self, profile_csv_list, ssl_profile_ref,
                                vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(NsxT type) csv rows
        :param ssl_profile_ref: Reference of ssl profile
        :param vs_ref: Name of VS
        :return: ssl profile name and skipped sttribute list
        """

        ssl_profile_name = self.get_name(ssl_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, ssl_profile_name, vs_ref, field_key='ssl_profile')
        return ssl_profile_name, skipped_list

    def update_vs_complexity_level(self, vs_csv_row, virtual_service):
        """
        This function defines that update complexity level of VS objects.
        if it has reference of VS Datascript or Http policies -> Advanced
        else
        level -> Basic
        :param vs_csv_row: csv row of VS
        :param virtual_service: dict of Virtual service
        :return: None
        """

        if ('http_policies' in virtual_service and
            virtual_service['http_policies']) or \
                ('vs_datascripts' in virtual_service and
                 virtual_service['vs_datascripts']):
            vs_csv_row['Complexity Level'] = conv_const.COMPLEXITY_ADVANCED
        else:
            vs_csv_row['Complexity Level'] = conv_const.COMPLEXITY_BASIC

    def remove_dup_key(self, obj_list):
        for obj in obj_list:
            obj.pop('dup_of', None)

    def check_for_duplicates(self, src_obj, obj_list, obj_type,
                             merge_object_mapping, ent_type, prefix, syslist):
        """
        Checks for duplicate objects except name and description values
        :param src_obj: Object to be checked for duplicate
        :param obj_list: List of oll objects to search in
        :return: Name of object for which given object is duplicate of
        """
        if not syslist:
            syslist = []
        src_cp = copy.deepcopy(src_obj)
        src_cp.pop("name")
        src_cp.pop("description", [])
        for obj in syslist:
            ob_cp = copy.deepcopy(obj)
            ob_cp.pop("name")
            ob_cp.pop("description", [])
            ob_cp.pop('url', [])
            ob_cp.pop('uuid', [])
            if src_cp.items() == ob_cp.items():
                return obj["name"], src_obj['name']
        for tmp_obj in obj_list:
            tmp_cp = copy.deepcopy(tmp_obj)
            tmp_cp.pop("name")
            tmp_cp.pop("description", [])
            dup_lst = tmp_cp.pop("dup_of", [tmp_obj["name"]])
            if src_cp.items() == tmp_cp.items():
                dup_lst.append(src_obj["name"])
                tmp_obj["dup_of"] = dup_lst
                old_name = tmp_obj['name']
                if tmp_obj["name"] in merge_object_mapping[obj_type].keys():
                    merge_object_mapping[obj_type]['no'] += 1
                    no = merge_object_mapping[obj_type]['no']
                    mid_name = ent_type and ('Merged-%s-%s-%s-%s' % (ent_type,
                                                                     obj_type, ran_str, str(no))) or (
                                       'Merged-%s-%s-%s' % (obj_type, ran_str, str(no)))
                    new_name = '%s-%s' % (prefix, mid_name) if prefix else \
                        mid_name
                    tmp_obj["name"] = new_name
                return tmp_obj["name"], old_name
        return None, None

    def upload_file(self, file_path):
        """
        Reads the given file and returns the UTF-8 string
        :param file_path: Path of file to read
        :return: UTF-8 string read from file
        """

        file_str = None
        # remove double quotes if available
        file_path = file_path.replace('"', '')

        if '/Common/' in file_path:
            file_path = file_path.replace('/Common/', '')
        if ':Common:' in file_path:
            file_path = file_path.replace(':Common:', '')
        try:
            with open(file_path, "r") as file_obj:
                file_str = file_obj.read()
                file_str = file_str  # .decode("utf-8")
        except UnicodeDecodeError:
            try:
                file_str = file_str  # .decode('latin-1')
            except:
                update_count('error')
                LOG.error("[UnicodeDecode] Error to read file %s" % file_path,
                          exc_info=True)
        except IOError:
            update_count('warning')
            LOG.warn("Cannot read file %s" % file_path)
        except:
            update_count('error')
            LOG.error("Error accessing file %s" % file_path, exc_info=True)
        return file_str

    def get_name(self, url):
        """
        This function defines that return name object from url
        :param url:
        :return: Name of object
        """
        parsed = urlparse(url)
        return parse_qs(parsed.query)['name'][0]

    def get_tenant_from_ref(self, url):
        """
        This function defines that return tenant from url
        :param url:
        :return: Name of tenant
        """
        parsed = urlparse(url)
        return parse_qs(parsed.query).get('tenant', ['admin'])[0]

    def get_obj_type_from_ref(self, url):
        return url.split('/api/')[1].split('/')[0].split('?')[0]

    def get_object_ref(self, object_name, object_type, tenant='admin',
                       cloud_name='Default-Cloud', prefix=None):
        """
        This function defines that to genarte object ref in the format of
        /api/object_type/?tenant=tenant_name&name=object_name&cloud=cloud_name
        :param object_name: Name of object
        :param object_type: Type of object
        :param tenant: Name of tenant
        :param cloud_name: Name of cloud
        :param prefix : Prefix for objects
        :return: Return generated object ref
        """
        # Added prefix for objects
        if prefix:
            object_name = prefix + '-' + object_name

        cloud_supported_types = ['pool', 'poolgroup', 'vsvip', 'vrfcontext',
                                 'serviceenginegroup', 'network']
        if not cloud_name:
            cloud_name = "Default-Cloud"

        if object_type == 'tenant':
            ref = '/api/tenant/?name=%s' % object_name
            if object_name not in tenants:
                tenants.append(object_name)
        elif object_type == 'cloud':
            ref = '/api/%s/?tenant=admin&name=%s' % (object_type, object_name)
        elif object_type == 'vrfcontext':
            ref = '/api/%s/?tenant=admin&name=%s&cloud=%s' % (
                object_type, object_name, cloud_name)
        elif object_type in cloud_supported_types:
            ref = '/api/%s/?tenant=%s&name=%s&cloud=%s' % (
                object_type, tenant, object_name, cloud_name)
        else:
            ref = '/api/%s/?tenant=%s&name=%s' % (
                object_type, tenant, object_name)
        # if cloud_name:
        #     ref += '&cloud=%s' % cloud_name
        return ref

    # Print iterations progress
    def print_progress_bar(self, iteration, total, msg, prefix='', suffix='',
                           decimals=1, length=50, fill='#', printEnd="\\"):
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent
            complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
        """
        percent = ("{0:." + str(decimals) + "f}"). \
            format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        if (iteration < total):
            print(f'\r{prefix} |{bar}| {percent}% {suffix}', end=printEnd)
        else:
            print(f'\r{prefix} |{bar}| {percent}% {suffix}')

    def validate_value(self, entity_names, prop_name, value, limit_data, obj,
                       valname):

        """
        :param entity_names: list of name of the avi entity/object
        :param prop_name: property name
        :param value: property value
        :param limit_data: data from generated yaml
        :param obj: obj name
        :return: validity and new value
        """
        valid = None
        new_value = value
        msgvar = valname and entity_names and '%s-->%s-->%s' % (valname,
                                                                '-->'.join(entity_names),
                                                                prop_name) or valname and '%s-->%s' \
                 % (valname, prop_name) or entity_names and '%s-->%s' % (
                     '-->'.join(entity_names), prop_name) or '%s' % prop_name
        for key, val in limit_data.items():
            pr = val.get(obj, {})
            if not pr:
                continue
            LOG.debug("Validating property '%s'", msgvar)
            p_key = self.get_to_prop(val, pr, entity_names, prop_name,
                                     limit_data)
            typ = p_key.get('py_type') if p_key else None
            if typ:
                break
        else:
            LOG.debug("Property '%s' is not present in generated yaml, reason "
                      "being the property doesn't have any attribute from the "
                      "list %s", msgvar, str(['default_value', 'range',
                                              'special_values', 'ref_type']))
            return None, None
        if new_value is not None:
            # commenting this since now in Python 3 strings are already in unicode format.
            if type(new_value) == str:
                new_value = new_value.encode()
            if type(new_value) == eval(typ) or (eval(typ) == int and
                                                str(new_value).isdigit()):
                special_value = p_key.get('special_values')
                if typ == 'int':
                    new_value = int(new_value)
                    if special_value and str(new_value) in special_value:
                        valid = True
                        LOG.debug("Special value %s is fine", str(new_value))
                    else:
                        lval = p_key.get('range')
                        if lval:
                            low, high = lval.split('-')
                            if new_value < int(low):
                                valid = False
                                new_value = int(low)
                            elif new_value > int(high):
                                valid = False
                                new_value = int(high)
                            else:
                                valid = True
                                LOG.debug("Value '%s' is fine", str(new_value))
                if typ == 'str':
                    if special_value and new_value in special_value:
                        valid = True
                        LOG.debug("Special value %s is fine", str(new_value))
                    else:
                        options = p_key.get('option_values')
                        if options and new_value not in options:
                            valid = False
                            new_value = p_key.get('default_value')
                        else:
                            valid = True
                            LOG.debug("Value '%s' is fine", str(new_value))
                if typ == 'bool':
                    if new_value not in (False, True, 'False', 'True'):
                        valid = False
                        new_value = p_key.get('default_value')
                        new_value = False if new_value == 'False' else True
                    else:
                        valid = True
                        LOG.debug("Value '%s' is fine", str(new_value))
            else:
                LOG.debug("Type of value '%s' doesn't match with type '%s' "
                          "defined", str(type(new_value)), typ)
                valid, new_value = None, None
        else:
            if p_key.get('required') == 'True':
                valid = False
                new_value = p_key.get('default_value')
                if typ == 'bool':
                    new_value = False if new_value == 'False' else True
                LOG.debug("Value is required hence, defaulting to value '%s'",
                          str(new_value))
            else:
                valid = True
                LOG.debug("Property '%s' not mandatory", msgvar)

        return valid, new_value

    def get_to_prop(self, val, pr, entity_names, prop_name, limit_data):

        for name in entity_names:
            if pr.get('properties', {}).get(name, {}):
                if pr['properties'][name].get('ref_type'):
                    ref = pr['properties'][name].get('ref_type')
                    vr = val.get(ref, {})
                    if not vr:
                        for k, v in limit_data.items():
                            if v.get(ref):
                                tr = v[ref]
                                return self.get_to_prop(v, tr, entity_names,
                                                        prop_name, limit_data)
                        else:
                            return
                    else:
                        return self.get_to_prop(val, vr, entity_names,
                                                prop_name, limit_data)
        p_key = pr.get('properties', {}).get(prop_name, {})
        if p_key:
            return p_key

    def validation(self, avi_config):

        """
        Validator function for all avi objects
        :param avi_config:
        :return:
        """

        LOG.debug("Starting Validation checks ... ")
        limit_data = None
        dir_path = os.path.abspath(os.path.dirname(__file__))
        if os.path.exists(dir_path + os.path.sep + 'pb_attributes.yaml'):
            with open(dir_path + os.path.sep + 'pb_attributes.yaml') as data:
                limit_data = yaml.safe_load(data)
        if limit_data:
            for obj, vals in avi_config.items():
                if obj != 'META' and vals:
                    for val in vals:
                        heir = []
                        LOG.debug("Validating %s of Object %s", val['name'],
                                  obj)
                        self.validate_prop(val, heir, limit_data, obj,
                                           val['name'])

    def validate_prop(self, dictval, heir, limit_data, obj, valname=None):

        for k, v in dictval.items():
            if k in ['tenant_ref', 'name', 'cloud_ref', 'health_monitor_refs',
                     'ssl_profile_ref', 'application_persistence_profile_ref',
                     'application_profile_ref', 'network_profile_ref',
                     'pki_profile_ref', 'pool_ref', 'pool_group_ref',
                     'http_policy_set_ref', 'ssl_key_and_certificate_refs',
                     'vsvip_ref', 'description']:
                msvar = valname and heir and '%s-->%s-->%s' % (valname,
                                                               '-->'.join(heir), k) or valname and '%s-->%s' % (
                            valname, k) or heir and '%s-->%s' % ('-->'.join(heir),
                                                                 k) or '%s' % k
                LOG.debug("Skipping validation checks for '%s'", msvar)
                continue
            else:
                if isinstance(v, list):
                    for listval in v:
                        if isinstance(listval, dict):
                            k not in heir and heir.append(k) or None
                            self.validate_prop(listval, heir, limit_data, obj,
                                               valname)
                            heir and heir.pop() or None
                        else:
                            mgvar = valname and heir and '%s-->%s-->%s' % (
                                valname, '-->'.join(heir), k) or valname \
                                    and '%s-->%s' % (valname, k) or heir and \
                                    '%s-->%s' % ('-->'.join(heir), k) or \
                                    '%s' % k
                            LOG.debug("Property '%s' has value as a list %s, "
                                      "not supported currently", mgvar, str(v))
                            # valid, val = self.validate_value(heir, k, listval,
                            # limit_data)
                            # if valid is False:
                            # LOG.debug("Correcting the value for %s" % k)
                            # dictval[k] = val
                elif isinstance(v, dict):
                    k not in heir and heir.append(k) or None
                    self.validate_prop(v, heir, limit_data, obj, valname)
                    heir and heir.pop() or None
                else:
                    valid, val = self.validate_value(heir, k, v, limit_data,
                                                     obj, valname)
                    if valid is False:
                        mvar = valname and heir and '%s-->%s-->%s' % (valname,
                                                                      '-->'.join(heir), k) or valname and '%s-->%s' % (
                                   valname, k) or heir and '%s-->%s' % (
                                   '-->'.join(heir), k) or '%s' % k
                        LOG.debug("Correcting the value for '%s' from '%s' to "
                                  "'%s'", mvar, str(v), str(val))
                        dictval[k] = val

    def check_certificate_expiry(self, input_dir, cert_file_name):
        cert_file_name = cert_file_name.replace(':Common:', '')
        cert_text = self.upload_file(input_dir + os.path.sep + cert_file_name)
        if cert_text:
            cert_date = crypto.load_certificate(crypto.FILETYPE_PEM,
                                                cert_text if type(cert_text) == str
                                                else cert_text.decode())
            expiry_date = datetime.strptime(cert_date.get_notAfter().decode('utf-8'),
                                            "%Y%m%d%H%M%SZ")
            present_date = datetime.now()
            if expiry_date < present_date:
                LOG.warning("Certificate %s is expired creating self "
                            "signed cert." % cert_file_name)
                return False
            else:
                return True
        else:
            LOG.warn('Cannot load cert %s to check expiry skipping the check' %
                     cert_file_name)
            update_count('warning')
            return True

    def make_graph(self, avi_config):
        """
        Filters vs and its references from full configuration
        :param avi_config: full configuration
        :return: graph
        """
        avi_graph = nx.DiGraph()
        avi_graph.add_node('AVI', type='Tree')
        for vs in avi_config.get('VirtualService',[]):
            name = vs['name']
            avi_graph.add_node(name, type='VS')
            avi_graph.add_edge('AVI', name)
            self.find_and_add_ne(vs, avi_config, avi_graph, name, depth=0)
        return avi_graph

    def find_and_add_ne(self, obj_dict, avi_config, avi_graph, vsname, depth):
        """
        Method to iterate in one object find references and add those to output
        :param obj_dict: Object to be iterated over
        :param avi_config: Full config
        :param avi_graph: avi graph
        :param vsname: name of object
        :param depth: Recursion depth to determine level in the vs reference
                      tree
        """

        for key in obj_dict:

            if (key.endswith('ref') and key not in ['cloud_ref', 'tenant_ref']) \
                    or key == 'ssl_profile_name':
                if not obj_dict[key]:
                    continue
                entity, name = self.get_name_and_entity(obj_dict[key])
                self.search_ne(entity, name, avi_graph, avi_config, depth,
                               vsname)
            elif key.endswith('refs'):
                for ref in obj_dict[key]:
                    entity, name = self.get_name_and_entity(ref)
                    self.search_ne(entity, name, avi_graph, avi_config, depth,
                                   vsname)
            elif isinstance(obj_dict[key], dict):
                self.find_and_add_ne(obj_dict[key], avi_config, avi_graph,
                                     vsname, depth)
            elif obj_dict[key] and isinstance(obj_dict[key], list) \
                    and isinstance(obj_dict[key][0], dict):
                for member in obj_dict[key]:
                    self.find_and_add_ne(member, avi_config, avi_graph, vsname,
                                         depth)
            elif key == 'hostname':
                sername = obj_dict[key]
                if avi_graph.has_node(sername):
                    node_type = [n[1]['type'] for n in list(
                        avi_graph.nodes().data()) if n[0] == sername]
                    node_type = '{}-{}'.format(node_type[0], 'Server')
                    avi_graph.add_node(sername, type=node_type)
                    avi_graph.add_edge(vsname, sername)
                else:
                    avi_graph.add_node(sername, type='Server')
                    avi_graph.add_edge(vsname, sername)
            elif key == 'name' and depth == 1 and '-rule-' in obj_dict[key]:
                rule_name = obj_dict[key]
                if avi_graph.has_node(rule_name):
                    node_type = [n[1]['type'] for n in list(
                        avi_graph.nodes().data()) if n[0] == rule_name]
                    node_type = '{}-{}'.format(node_type[0], 'Rule')
                    avi_graph.add_node(rule_name, type=node_type)
                    avi_graph.add_edge(vsname, rule_name)
                else:
                    avi_graph.add_node(rule_name, type='Rule')
                    avi_graph.add_edge(vsname, rule_name)
        return

    def search_ne(self, entity, name, avi_graph, avi_config, depth, vsname):
        """
        Method to search referenced object
        :param entity: object type
        :param name: object name
        :param avi_graph: avi graph
        :param avi_config: full config
        :param depth: Recursion depth to determine level in the vs reference
                      tree
        :param vsname: name of VS
        """
        path_key_map = self.get_path_key_map()
        avi_conf_key = path_key_map[entity]
        found_obj_list = avi_config.get(avi_conf_key)
        found_obj = [obj for obj in found_obj_list if obj['name'] == name] if \
            found_obj_list else []
        if found_obj:
            found_obj = found_obj[0]
            if avi_graph.has_node(name):
                nod_type = [node[1]['type'] for node in list(
                    avi_graph.nodes().data()) if node[0] == name]
                nod_type = '{}-{}'.format(nod_type[0], path_key_map[entity])
                avi_graph.add_node(name, type=nod_type)
                avi_graph.add_edge(vsname, name)
            else:
                avi_graph.add_node(name, type=path_key_map[entity])
                avi_graph.add_edge(vsname, name)
        elif entity in ['applicationprofile', 'networkprofile', 'healthmonitor',
                        'sslkeyandcertificate', 'sslprofile',
                        'applicationpersistenceprofile']:
            if str.startswith(str(name), 'System-'):
                return
        else:
            update_count()
            LOG.warning('Reference not found for %s with name %s' % (
                entity, name))
            return
        depth += 1
        new_name=""
        if found_obj:
            new_name = found_obj.get('name')
        if new_name:
            self.find_and_add_ne(found_obj, avi_config, avi_graph, new_name,
                                 depth)

    def get_project_path(self):
        return os.path.abspath(os.path.dirname(__file__))

    def get_name_and_entity(self, url):
        """
        Parses reference string to extract object type and
        :param url: reference url to be parsed
        :return: entity and object name
        """
        parsed = urlparse(url)
        return parsed.path.split('/')[2], \
               parse_qs(parsed.query)['name'][0]

    def get_path_key_map(self):
        yml_file = os.path.join(self.get_project_path(),
                                './common/avi_resource_types.yaml')
        yml_data = yaml.load(open(yml_file, 'r'), Loader=yaml.Loader)
        # Converts avi object types to avi resource types
        data_lower_case = map(lambda x: x.lower(),
                              yml_data['avi_resource_types'])
        # Generates AVI resource types to avi object type mapping
        # in form of dictionary.
        return dict(zip(data_lower_case, yml_data['avi_resource_types']))

    def clone_app_profile_for_vs(self, app_prof_ref, app_prof_obj, vs_name,
                                 tenant, avi_config):
        app_prof_name = self.get_name(app_prof_ref)
        new_app_profile = copy.deepcopy(app_prof_obj)
        new_profile_name = '%s-%s' % (app_prof_name, vs_name)
        new_app_profile['name'] = new_profile_name
        avi_config['ApplicationProfile'].append(new_app_profile)
        new_ref = self.get_object_ref(
            new_profile_name, 'applicationprofile', tenant)
        return new_ref

    def add_complete_conv_status(self, output_dir, avi_config, report_name,
                                 vs_level_status):

        global csv_writer_dict_list
        global ptotal_count
        for status in conv_const.STATUS_LIST:
            status_list = [row for row in csv_writer_dict_list if
                           row['Status'] == status]
            print('%s: %s' % (status, len(status_list)))
        print("Writing Excel Sheet For Converted Configuration...")
        ptotal_count = ptotal_count + len(csv_writer_dict_list)
        if vs_level_status:
            self.vs_per_skipped_setting_for_references(avi_config)
            self.correct_vs_ref(avi_config)
        else:
            # Update the complexity level of VS as Basic or Advanced
            self.vs_complexity_level()
        self.write_status_report_and_pivot_table_in_xlsx(
            output_dir, report_name, vs_level_status)

    def write_status_report_and_pivot_table_in_xlsx(
            self, output_dir, report_name, vs_level_status):
        """
        This function defines that add status sheet and pivot table sheet in
        xlsx format
        :param output_dir: Path of output directory
        :param report_name: filename to write report
        :param vs_level_status: Flag to include VS wise detailed status or not
        :return: None
        """
        global ppcount
        global ptotal_count
        # List of fieldnames for headers
        if vs_level_status:
            fieldnames = ['F5 type', 'F5 SubType', 'F5 ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable', 'User Ignored',
                          'Skipped for defaults', 'Complexity Level',
                          'VS Reference', 'Overall skipped settings',
                          'Avi Object', 'Needs Review']
        else:
            fieldnames = ['F5 type', 'F5 SubType', 'F5 ID', 'Status',
                          'Skipped settings', 'Indirect mapping',
                          'Not Applicable',
                          'User Ignored', 'Skipped for defaults',
                          'Complexity Level', 'Avi Object', 'Needs Review']

        # xlsx workbook
        report_path = output_dir + os.path.sep + "%s-ConversionStatus.xlsx" % \
                      report_name
        status_wb = Workbook(report_path)
        # xlsx worksheet
        status_ws = status_wb.add_worksheet("Status Sheet")
        # Lock the first row of xls report.
        status_ws.freeze_panes(1, 0)
        first_row = 0
        for header in fieldnames:
            col = fieldnames.index(header)
            status_ws.write(first_row, col, header)
        row = 1
        for row_data in csv_writer_dict_list:
            ppcount += 1
            for _key, _value in row_data.items():
                col = fieldnames.index(_key)
                status_ws.write(row, col, _value)
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
            row += 1
        status_wb.close()
        # create dataframe for row list
        df = pandas.DataFrame(csv_writer_dict_list, columns=fieldnames)
        # create pivot table using pandas
        pivot_table = \
            pandas.pivot_table(df, index=["Status", "F5 type", "F5 SubType"],
                               values=[], aggfunc=[len], fill_value=0)
        # create dataframe for pivot table using pandas
        pivot_df = pandas.DataFrame(pivot_table)
        main_book = \
            load_workbook(report_path)
        main_writer = pandas.ExcelWriter(report_path, engine='openpyxl')
        main_writer.book = main_book
        # Add pivot table in Pivot sheet
        pivot_df.to_excel(main_writer, 'Pivot Sheet')
        main_writer.save()

    def vs_complexity_level(self):
        """
        This method calculate the complexity of vs.
        :return:
        """
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['F5 type'] == 'virtual']
        for vs_csv_object in vs_csv_objects:
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)

    def correct_vs_ref(self, avi_config):
        """
        This method corrects the reference of VS to different objects
        :param avi_config: avi configuration dict
        :return:
        """
        global csv_writer_dict_list
        avi_graph = self.make_graph(avi_config)
        csv_dict_sub = [row for row in csv_writer_dict_list if row[
                        'F5 type'] != 'virtual' and row['Status'] in
                        (conv_const.STATUS_PARTIAL,
                         conv_const.STATUS_SUCCESSFUL)]
        for dict_row in csv_dict_sub:
            obj = dict_row['Avi Object']
            vs = []
            if obj.startswith('{'):
                obj = eval(obj)
                for key in obj:
                    for objs in obj[key]:
                        self.add_vs_ref(objs, avi_graph, vs)
            elif obj.startswith('['):
                obj = eval(obj)
                for objs in obj:
                    for key in objs:
                        objval = objs[key]
                        self.add_vs_ref(objval, avi_graph, vs)
            if vs:
                dict_row['VS Reference'] = str(list(set(vs)))
            else:
                dict_row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def add_vs_ref(self, obj, avi_graph, vs):
        """
        Helper method for adding vs ref
        :param obj: object
        :param avi_graph: avi graph
        :param vs: VS list
        :return:
        """
        tmplist = []
        if isinstance(obj, str) and obj.startswith('Duplicate of'):
            obj_name = None
            LOG.debug("Object has merged: %s" % obj)
        else:
            obj_name = obj.get('name', obj.get('hostname'))
        if obj_name:
            if avi_graph.has_node(obj_name):
                LOG.debug("Checked predecessor for %s", obj_name)
                predecessor = list(avi_graph.predecessors(obj_name))
                if predecessor:
                    self.get_predecessor(predecessor, avi_graph, vs, tmplist)
            else:
                LOG.debug("Object %s may be merged or orphaned", obj_name)

    def get_predecessor(self, predecessor, avi_graph, vs, tmplist):
        """
        This method gets the predecessor of the object
        :param predecessor: predecessor list
        :param avi_graph: avi graph
        :param vs: VS list
        :param tmplist: temporary list of objects for which predecessors
                        are already evaluated
        :return:
        """
        if len(predecessor) > 1:
            for node in predecessor:
                if node in tmplist:
                    continue
                nodelist = [node]
                self.get_predecessor(nodelist, avi_graph, vs, tmplist)
        elif len(predecessor):
            node_obj = [nod for nod in list(avi_graph.nodes().data()) if
                        nod[0] == predecessor[0]]
            if node_obj and (node_obj[0][1]['type'] == 'VS' or 'VS' in node_obj[
                0][1]['type']):
                LOG.debug("Predecessor %s found", predecessor[0])
                vs.extend(predecessor)
            else:
                tmplist.extend(predecessor)
                LOG.debug("Checked predecessor for %s", predecessor[0])
                nodelist = list(avi_graph.predecessors(predecessor[0]))
                self.get_predecessor(nodelist, avi_graph, vs, tmplist)
        else:
            LOG.debug("No more predecessor")

    def vs_per_skipped_setting_for_references(self, avi_config):
        """
        This functions defines that Add the skipped setting per VS CSV row
        :param avi_config: this method use avi_config for checking vs skipped
        :return: None
        """

        # Get the count of vs fully migrated
        global fully_migrated
        global ptotal_count
        global ppcount
        fully_migrated = 0
        # Get the VS object list which is having status successful and partial.
        vs_csv_objects = [row for row in csv_writer_dict_list
                          if row['Status'] in [conv_const.STATUS_PARTIAL,
                                               conv_const.STATUS_SUCCESSFUL]
                          and row['F5 type'] == 'virtual']
        # Get the list of csv rows which has profile as F5 Type
        profile_csv_list = self.get_csv_object_list(
            csv_writer_dict_list, ['profile'])
        ptotal_count = ptotal_count + len(vs_csv_objects)
        for vs_csv_object in vs_csv_objects:
            ppcount += 1
            skipped_setting = {}
            virtual_service = self.format_string_to_json(
                vs_csv_object['Avi Object'])
            # Update the complexity level of VS as Basic or Advanced
            self.update_vs_complexity_level(vs_csv_object, virtual_service)
            vs_ref = virtual_service['name']
            repls = ('[', ''), (']', '')
            # Get list of skipped setting attributes
            skipped_setting_csv = reduce(lambda a, kv: a.replace(*kv), repls,
                                         vs_csv_object['Skipped settings'])
            if skipped_setting_csv:
                skipped_setting['virtual_service'] = [skipped_setting_csv]
            # Get the skipped list for ssl key and cert
            if 'ssl_key_and_certificate_refs' in virtual_service:
                for ssl_key_and_certificate_ref in \
                        virtual_service['ssl_key_and_certificate_refs']:
                    ssl_key_cert = self.get_name(ssl_key_and_certificate_ref)
                    ssl_kc_skip = self.get_csv_skipped_list(
                        profile_csv_list, ssl_key_cert, vs_ref,
                        field_key='ssl_cert_key')
                    if ssl_kc_skip:
                        skipped_setting['ssl cert key'] = {}
                        skipped_setting['ssl cert key']['name'] = ssl_key_cert
                        skipped_setting['ssl cert key'][
                            'skipped_list'] = ssl_kc_skip

            # Get the skipped list for ssl profile name.
            # Changed ssl profile name to ssl profile ref.
            if 'ssl_profile_ref' in virtual_service:
                name, skipped = self.get_ssl_profile_skipped(
                    profile_csv_list, virtual_service['ssl_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['ssl profile'] = {}
                    skipped_setting['ssl profile']['name'] = name
                    skipped_setting['ssl profile']['skipped_list'] = skipped
            # Get the skipped list for pool group.
            if 'pool_group_ref' in virtual_service:
                pool_group_name = self.get_name(
                    virtual_service['pool_group_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                pool_group_skipped_settings = self.get_pool_skipped_list(
                    avi_config, pool_group_name, csv_pool_rows,
                    csv_writer_dict_list, vs_ref, profile_csv_list)
                if pool_group_skipped_settings:
                    skipped_setting['Pool Group'] = pool_group_skipped_settings
            # Get the skipped list for pool.
            if 'pool_ref' in virtual_service:
                pool_skipped_settings = {'pools': []}
                pool_name = self.get_name(virtual_service['pool_ref'])
                csv_pool_rows = self.get_csv_object_list(csv_writer_dict_list,
                                                         ['pool'])
                self.get_skipped_pool(
                    avi_config, pool_name, csv_pool_rows, csv_writer_dict_list,
                    vs_ref, profile_csv_list, pool_skipped_settings)
                if pool_skipped_settings['pools']:
                    skipped_setting['Pool'] = pool_skipped_settings
            # Get the skipepd list for http policy.
            if 'http_policies' in virtual_service:
                policy_csv_list = self.get_csv_object_list(
                    csv_writer_dict_list, ['policy', 'profile'])
                for http_ref in virtual_service['http_policies']:
                    policy_set_name, skipped_list = self.get_policy_set_skipped(
                        policy_csv_list, http_ref['http_policy_set_ref'],
                        vs_ref)
                    if skipped_list:
                        skipped_setting['Httppolicy'] = {}
                        skipped_setting['Httppolicy']['name'] = policy_set_name
                        skipped_setting['Httppolicy'][
                            'skipped_list'] = skipped_list
                    # Get the http policy name
                    pool_csv_rows = \
                        self.get_csv_object_list(csv_writer_dict_list, ['pool'])
                    for each_http_policy in avi_config['HTTPPolicySet']:
                        if each_http_policy['name'] == policy_set_name and 'http_request_policy' in each_http_policy:
                            for http_req in each_http_policy[
                              'http_request_policy']['rules']:
                                if http_req.get('switching_action', {}):
                                    self.get_skip_pools_policy(
                                        policy_set_name, http_req,
                                        avi_config, pool_csv_rows, vs_ref,
                                        profile_csv_list, skipped_setting)

            # # Get the skipped list for application_profile_ref.
            if 'application_profile_ref' in virtual_service and 'admin:System' \
                    not in virtual_service['application_profile_ref']:
                name, skipped = self.get_application_profile_skipped(
                    profile_csv_list,
                    virtual_service['application_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Application profile'] = {}
                    skipped_setting['Application profile'][
                        'name'] = name
                    skipped_setting['Application profile'][
                        'skipped_list'] = skipped
            # # Get the skipped list for network profile ref.
            if 'network_profile_ref' in virtual_service and 'admin:System' \
                    not in virtual_service['network_profile_ref']:
                name, skipped = self.get_network_profile_skipped(
                    profile_csv_list, virtual_service['network_profile_ref'],
                    vs_ref)
                if skipped:
                    skipped_setting['Network profile'] = {}
                    skipped_setting['Network profile'][
                        'name'] = name
                    skipped_setting['Network profile'][
                        'skipped_list'] = skipped
            # Update overall skipped setting of VS csv row
            if skipped_setting:
                vs_csv_object.update(
                    {'Overall skipped settings': str(skipped_setting)})
            else:
                vs_csv_object.update(
                    {'Overall skipped settings': "FULLY MIGRATION"})
                fully_migrated += 1
            # Added call for progress function.
            msg = "excel sheet conversion started..."
            self.print_progress_bar(ppcount, ptotal_count, msg,
                                    prefix='Progress', suffix='')
        csv_objects = [row for row in csv_writer_dict_list
                       if row['Status'] in [
                           conv_const.STATUS_PARTIAL,
                           conv_const.STATUS_SUCCESSFUL]
                       and row['F5 type'] != 'virtual']

        # Update the vs reference not in used if objects are not attached to
        # VS directly or indirectly
        for row in csv_objects:
            if 'VS Reference' not in row or row['VS Reference'] == '':
                row['VS Reference'] = conv_const.STATUS_NOT_IN_USE

    def get_application_profile_skipped(self, profile_csv_list, app_profile_ref,
                                        vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param app_profile_ref: Reference of application profile
        :param vs_ref: Name of VS
        :return: application profile name and skipped sttribute list
        """

        app_profile_name = self.get_name(app_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, app_profile_name, vs_ref, field_key='app_profile')
        return app_profile_name, skipped_list

    def get_network_profile_skipped(self, profile_csv_list, network_profile_ref,
                                    vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param network_profile_ref: Reference of Network profile
        :param vs_ref: Name of VS
        :return: network profile name and skipped sttribute list
        """

        network_profile_name = self.get_name(network_profile_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, network_profile_name, vs_ref,
            field_key='network_profile')
        return network_profile_name, skipped_list

    def get_policy_set_skipped(self, profile_csv_list, policy_set_ref, vs_ref):
        """
        This functions defines that get the skipped list of CSV row
        :param profile_csv_list: List of profile(F5 type) csv rows
        :param policy_set_ref: Reference of policy set
        :param vs_ref: Name of VS
        :return: policy set name and skipped sttribute list
        """

        policy_set_name = self.get_name(policy_set_ref)
        skipped_list = self.get_csv_skipped_list(
            profile_csv_list, policy_set_name, vs_ref, field_key='policy_set')
        return policy_set_name, skipped_list

    def get_skip_pools_policy(self, policy_set_name, http_req, avi_config,
                              pool_csv_rows, vs_ref, profile_csv_list,
                              skipped_setting):
        if http_req['switching_action'].get('pool_group_ref'):
            pool_group_name = self.get_name(http_req['switching_action']
                                            ['pool_group_ref'])
            pool_group_skipped_settings = self.get_pool_skipped_list(
                avi_config, pool_group_name, pool_csv_rows,
                csv_writer_dict_list, vs_ref, profile_csv_list)
            if pool_group_skipped_settings:
                if 'Httppolicy' not in skipped_setting:
                    skipped_setting['Httppolicy'] = {}
                    skipped_setting['Httppolicy']['name'] = policy_set_name
                skipped_setting['Httppolicy']['Pool Group'] =\
                    pool_group_skipped_settings
        elif http_req['switching_action'].get('pool_ref'):
            pool_name = self.get_name(http_req['switching_action']['pool_ref'])
            pool_skipped_settings = {'pools': []}
            self.get_skipped_pool(avi_config, pool_name, pool_csv_rows,
                                  csv_writer_dict_list, vs_ref,
                                  profile_csv_list, pool_skipped_settings)
            if pool_skipped_settings['pools']:
                if 'Httppolicy' not in skipped_setting:
                    skipped_setting['Httppolicy'] = {}
                    skipped_setting['Httppolicy']['name'] = policy_set_name
                skipped_setting['Httppolicy']['Pool'] = pool_skipped_settings

    def update_skip_duplicates(self, obj, obj_list, obj_type, converted_objs,
                               name, default_profile_name, merge_object_mapping,
                               ent_type, prefix, syslist):

        """
        Merge duplicate profiles
        :param obj: Source object to find duplicates for
        :param obj_list: List of object to search duplicates in
        :param obj_type: Type of object to add in converted_objs status
        :param converted_objs: Converted avi object or merged object name
        :param name: Name of the object
        :param default_profile_name : Name of root parent default profile
        :param merge_object_mapping: merged object mappings
        :param ent_type: Entity type
        :param prefix: object name prefix
        :param syslist: System object list
        :return:
        """
        dup_of = None
        if isinstance(merge_object_mapping, dict):
            merge_object_mapping[obj_type].update({name: name})
        # root default profiles are skipped for merging
        if not name == default_profile_name or obj_type == 'ssl_profile':
            dup_of, old_name = \
                self.check_for_duplicates(obj, obj_list, obj_type,
                                          merge_object_mapping, ent_type,
                                          prefix, syslist)
        if dup_of:
            converted_objs.append({obj_type: "Duplicate of %s" % dup_of})
            LOG.info(
                "Duplicate profiles: %s merged in %s" % (obj['name'], dup_of))
            if isinstance(merge_object_mapping, dict):
                if old_name in merge_object_mapping[obj_type].keys():
                    merge_object_mapping[obj_type].update({old_name: dup_of})
                merge_object_mapping[obj_type].update({name: dup_of})
        else:
            obj_list.append(obj)
            converted_objs.append({obj_type: obj})
